from datetime import datetime
import openpyxl
import pyodbc
import pandas
from pandas.io import sql
from openpyxl.utils.dataframe import dataframe_to_rows
import getpass

username = getpass.getpass("Enter username: ")
password = getpass.getpass("Enter password: ")

con = pyodbc.connect("Driver={NetezzaSQL};"
                     "Server=netezza-ew-con1;"
                    "Database=adhoc_usage;"
                    "uid="+username+";pwd="+password+";")

wb = openpyxl.load_workbook('U:\RDSAnalyticsMIS\PERS Ryan\Retail Heatmap\Heatmap Review Template.xlsx')


ws=wb["Data"]
#Query1
sql_query1 = pandas.read_sql_query(
    """
	## SQL query removed due to company data privacy.
;""", con)

df1=pandas.DataFrame(sql_query1, columns = ['clndr_yr_nbr','clndr_qtr_unq_id','clndr_mth_id','cd_ins','cd_outs','cd_nna','d_i','ms_nna','acat_in','acat_out','net_acat','schw_acat_in','schw_acat_out','net_schw_acat','acat_in_ex_schw','acat_out_ex_schw','net_acat_ex_schw','acat_ratio','acat_ratio_ex_schw','cd_ins_ex_schw','cd_outs_ex_schw','cd_nna_ex_schw','ms_nna_ex_schw','total_darts','tos_darts', 'tda_mobile_darts', 'mobile_trader_darts', 'web_darts', 'tdax_darts','other_darts'])

rows1=dataframe_to_rows(df1,index=False,header = True)

for r_idx, row in enumerate(rows1,1):
   for c_idx, value in enumerate(row,1):
        ws.cell(row=r_idx, column=c_idx, value = value)

## Query 2 that pulls all trade data for each quarter and month
ws2=wb["Data"]

sql_query2 = pandas.read_sql_query(
    """
		## SQL query removed due to company data privacy.
;""", con)

df2=pandas.DataFrame(sql_query2, columns = ['clndr_yr_nbr','clndr_qtr_unq_id','clndr_mth_id','security_type','total_darts'])

rows2=dataframe_to_rows(df2,index=False,header = True)

for r_idx, row in enumerate(rows2,1):
   for c_idx, value in enumerate(row,1):
        ws2.cell(row=r_idx, column=c_idx+31, value = value)

## Query 3 that pulls the number of trade days we have seen over the quarter/month
ws3=wb["Summary"]

sql_query3 = pandas.read_sql_query(
    """
		## SQL query removed due to company data privacy.
;""", con)

df3=pandas.DataFrame(sql_query3, columns = ['clndr_yr_nbr','clndr_qtr_unq_id','clndr_mth_id','trade_days'])

rows3=dataframe_to_rows(df3,index=False,header = True)

for r_idx, row in enumerate(rows3,1):
   for c_idx, value in enumerate(row,1):
        ws3.cell(row=r_idx, column=c_idx+7, value = value)

curr_year = int(input("What year are you verifying?  "))
curr_quarter = int(input("What CY quarter are you verifying?  "))
curr_month = int(input("What CY month are you verifying?  "))

ws3.cell(row=2, column=2, value=curr_quarter)
ws3.cell(row=3, column=2, value=curr_month)
ws3.cell(row=4, column=2, value=curr_year)

datestring = datetime.strftime(datetime.now(), ' %Y_%m_%d')

wb.save('U:\RDSAnalyticsMIS\PERS Ryan\Retail Heatmap\Heatmap Review for '+ datestring +' Heatmap.xlsx')

print("Heatmap data was successfully pulled and ready to be validated!")