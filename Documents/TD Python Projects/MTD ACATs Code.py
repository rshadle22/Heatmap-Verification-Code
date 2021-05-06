from datetime import datetime
import openpyxl
import pyodbc
import pandas
from pandas.io import sql
from openpyxl.utils.dataframe import dataframe_to_rows
import getpass

username = getpass.getpass("Enter username: ")
password = getpass.getpass("Enter password: ")

con = pyodbc.connect("Driver={NetezzaSQL};"
                     "Server=netezza-ew-con1;"
                    "Database=adhoc_usage;"
                    "uid="+username+";pwd="+password+";")

wb = openpyxl.load_workbook('U:\RDSAnalyticsMIS\PERS Ryan\ACAT Work\MTD ACATs\MTD ACATs Template.xlsx')


ws=wb["Monthly ACATs"]
#Query used to pull all MTD ACATs
sql_query = pandas.read_sql_query(
    """
    	## SQL query removed due to company data privacy.
;""", con)

df=pandas.DataFrame(sql_query, columns = ['clndr_mth_id','comp_type','brokerage','acats_in','acats_out'])

rows=dataframe_to_rows(df,index=False,header = True)

for r_idx, row in enumerate(rows,1):
   for c_idx, value in enumerate(row,1):
        ws.cell(row=r_idx, column=c_idx, value = value)

curr_month = int(input("What is the current month you are comparing?  "))
last_month = int(input("What is the last month you are comparing to?  "))

ws.cell(row=1, column=9, value=curr_month)
ws.cell(row=1, column=14, value=last_month)

datestring = datetime.strftime(datetime.now(), ' %Y_%m_%d')

wb.save('U:\RDSAnalyticsMIS\PERS Ryan\ACAT Work\MTD ACATs\MTD ACATs Updated - '+ datestring +'.xlsx')

print("The MTD ACAT information is up to date and ready!")